"""RAG Manager - 知识库管理与检索增强生成"""

import os
import json
import uuid
import time
from typing import List, Optional, Dict, Tuple

import faiss
import numpy as np
from langchain_text_splitters import RecursiveCharacterTextSplitter

from embeddings import DashScopeEmbeddings

# 文档解析
try:
    import fitz  # PyMuPDF
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


class RagManager:
    """RAG 知识库管理器（FAISS + DashScope Embedding + Qwen LLM）"""

    def __init__(self, api_key: str, base_dir: str = None):
        self.api_key = api_key
        self.base_dir = base_dir or os.path.dirname(os.path.abspath(__file__))
        self.docs_dir = os.path.join(self.base_dir, 'data', 'docs')
        self.index_dir = os.path.join(self.base_dir, 'data', 'index')
        self.index_path = os.path.join(self.index_dir, 'faiss_index.bin')
        self.meta_path = os.path.join(self.index_dir, 'meta.json')

        os.makedirs(self.docs_dir, exist_ok=True)
        os.makedirs(self.index_dir, exist_ok=True)

        self.embedder = DashScopeEmbeddings(api_key=api_key)
        self.splitter = RecursiveCharacterTextSplitter(
            chunk_size=800,
            chunk_overlap=150,
            separators=["\n\n", "\n", "。", "！", "？", "；", "，", " ", ""]
        )

        self.chunks: List[str] = []
        self.metadata: List[Dict] = []
        self.index: Optional[faiss.IndexFlatIP] = None
        self.embedding_dim = 1024  # text-embedding-v3 维度

        self._load_index()

    # ============ 文档解析 ============

    def _parse_pdf(self, filepath: str) -> str:
        if not HAS_PDF:
            return ""
        doc = fitz.open(filepath)
        texts = []
        for page in doc:
            texts.append(page.get_text())
        return "\n".join(texts)

    def _parse_docx(self, filepath: str) -> str:
        if not HAS_DOCX:
            return ""
        doc = DocxDocument(filepath)
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

    def _parse_xlsx(self, filepath: str) -> str:
        if not HAS_XLSX:
            return ""
        wb = openpyxl.load_workbook(filepath, data_only=True)
        texts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            texts.append(f"## 工作表: {sheet}")
            for row in ws.iter_rows(values_only=True):
                texts.append(" | ".join(str(c) for c in row if c is not None))
        return "\n".join(texts)

    def _parse_txt(self, filepath: str) -> str:
        for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    return f.read()
            except (UnicodeDecodeError, UnicodeError):
                continue
        return ""

    def parse_document(self, filepath: str) -> str:
        """解析各种格式的文档"""
        ext = os.path.splitext(filepath)[1].lower()
        parsers = {
            '.pdf': self._parse_pdf,
            '.docx': self._parse_docx,
            '.xlsx': self._parse_xlsx,
            '.xls': self._parse_xlsx,
            '.txt': self._parse_txt,
            '.md': self._parse_txt,
            '.csv': self._parse_txt,
            '.log': self._parse_txt,
        }
        parser = parsers.get(ext)
        if parser is None:
            return ""
        return parser(filepath)

    # ============ 索引管理 ============

    def _load_index(self):
        """从磁盘加载索引"""
        if os.path.exists(self.index_path) and os.path.exists(self.meta_path):
            self.index = faiss.read_index(self.index_path)
            with open(self.meta_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.chunks = data.get('chunks', [])
            self.metadata = data.get('metadata', [])
            self.embedding_dim = data.get('dim', 1024)

    def _save_index(self):
        """保存索引到磁盘"""
        if self.index is not None and len(self.chunks) > 0:
            faiss.write_index(self.index, self.index_path)
            with open(self.meta_path, 'w', encoding='utf-8') as f:
                json.dump({
                    'chunks': self.chunks,
                    'metadata': self.metadata,
                    'dim': self.embedding_dim,
                }, f, ensure_ascii=False, indent=2)

    # ============ 文档操作 ============

    def add_document(self, filepath: str) -> Dict:
        """添加文档到知识库"""
        filename = os.path.basename(filepath)
        content = self.parse_document(filepath)

        if not content or len(content.strip()) < 10:
            return {'success': False, 'error': '文档内容为空或无法解析'}

        # 切分文本
        new_chunks = self.splitter.split_text(content)
        if not new_chunks:
            return {'success': False, 'error': '文档无法切分为有效块'}

        doc_id = f"doc_{uuid.uuid4().hex[:8]}"

        # 生成嵌入
        embeddings = self.embedder.embed_documents(new_chunks)
        if not embeddings:
            return {'success': False, 'error': '嵌入生成失败'}

        self.embedding_dim = len(embeddings[0])

        # 构建元数据
        for i, chunk in enumerate(new_chunks):
            self.chunks.append(chunk)
            self.metadata.append({
                'doc_id': doc_id,
                'filename': filename,
                'chunk_index': i,
                'filepath': filepath,
            })

        # 更新 FAISS 索引
        vectors = np.array(embeddings, dtype=np.float32)
        # 归一化用于内积检索（余弦相似度）
        faiss.normalize_L2(vectors)

        if self.index is None:
            self.index = faiss.IndexFlatIP(self.embedding_dim)

        self.index.add(vectors)
        self._save_index()

        return {
            'success': True,
            'doc_id': doc_id,
            'filename': filename,
            'chunk_count': len(new_chunks),
        }

    def remove_document(self, doc_id: str) -> Dict:
        """移除文档（重建索引）"""
        new_chunks = []
        new_metadata = []
        keep_indices = []

        for i, meta in enumerate(self.metadata):
            if meta.get('doc_id') != doc_id:
                new_chunks.append(self.chunks[i])
                new_metadata.append(meta)
                keep_indices.append(i)

        if len(keep_indices) == len(self.chunks):
            return {'success': False, 'error': '未找到该文档'}

        # 重建索引
        if new_chunks:
            embeddings = self.embedder.embed_documents(new_chunks)
            vectors = np.array(embeddings, dtype=np.float32)
            faiss.normalize_L2(vectors)
            self.index = faiss.IndexFlatIP(self.embedding_dim)
            self.index.add(vectors)
        else:
            self.index = None
            if os.path.exists(self.index_path):
                os.remove(self.index_path)

        self.chunks = new_chunks
        self.metadata = new_metadata
        self._save_index()

        return {'success': True, 'remaining_docs': len(new_chunks)}

    # ============ 检索 ============

    def search(self, query: str, top_k: int = 5) -> List[Dict]:
        """检索相关文档块"""
        if self.index is None or len(self.chunks) == 0:
            return []

        query_vec = self.embedder.embed_query(query)
        query_vector = np.array([query_vec], dtype=np.float32)
        faiss.normalize_L2(query_vector)

        k = min(top_k, len(self.chunks))
        scores, indices = self.index.search(query_vector, k)

        results = []
        for i, idx in enumerate(indices[0]):
            if idx < len(self.chunks):
                results.append({
                    'content': self.chunks[idx],
                    'score': float(scores[0][i]),
                    'metadata': self.metadata[idx],
                })

        return results

    # ============ 知识库信息 ============

    def get_stats(self) -> Dict:
        """获取知识库统计信息"""
        docs = {}
        for meta in self.metadata:
            doc_id = meta.get('doc_id')
            if doc_id not in docs:
                docs[doc_id] = {
                    'doc_id': doc_id,
                    'filename': meta.get('filename'),
                    'chunk_count': 0,
                }
            docs[doc_id]['chunk_count'] += 1

        return {
            'total_chunks': len(self.chunks),
            'total_documents': len(docs),
            'documents': list(docs.values()),
        }

    def get_all_docs(self) -> List[Dict]:
        """获取所有文档列表"""
        docs = {}
        for meta in self.metadata:
            doc_id = meta.get('doc_id')
            if doc_id not in docs:
                docs[doc_id] = {
                    'doc_id': doc_id,
                    'filename': meta.get('filename'),
                    'chunk_count': 0,
                }
            docs[doc_id]['chunk_count'] += 1
        return list(docs.values())
